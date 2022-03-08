import psycopg2
import xlsxwriter
import openpyxl

hostname = 'localhost'
db = 'python-sql-assignment'
username = 'postgres'
pwd = '17JE003089'
port_id = 5432


class Assignment:
    # Method to list employee numbers, names, and their managers
    def list_emp_details(self, conn, workbook):
        query1 = "select empno, ename, mgr from emp"  # SQL query to list all employee details
        cur = conn.cursor()
        cur.execute(query1)
        rows = cur.fetchall()  # to execute the query using given cursor

        worksheet1 = workbook.add_worksheet("q1")  # Creating a new Worksheet in given xlsx workbook
        # print(worksheet1)

        rowIdx = 2

        # Creating the Column name
        worksheet1.write('A1', 'employee number')
        worksheet1.write('B1', 'names')
        worksheet1.write('C1', 'managers')

        print()
        print("list_emp_details")
        print()

        # Writing in xlsx worksheet
        for r in rows:
            worksheet1.write('A' + str(rowIdx), r[0])
            worksheet1.write('B' + str(rowIdx), r[1])
            worksheet1.write('C' + str(rowIdx), r[2])
            rowIdx += 1
            print(r)

        cur.close()

    # List the Total Cpmensation given till his/her last date
    def list_total_compensation(self, conn, workbook):
        # Created the query using CASE statement in which if there hiring date is NULL then we are subtracting with
        # current date else subtracting with leaving date.
        query1 = '''select
        	emp.ename,
        	emp.empno,
        	dept.dname,
         	emp.sal,
        	case
        		when jobhist.enddate is null then extract(year from age(CURRENT_TIMESTAMP, jobhist.startdate)) * 12 +
        		extract(month from age(CURRENT_TIMESTAMP, jobhist.startdate))
        		else extract(year from age(jobhist.enddate, jobhist.startdate)) * 12 +
        		extract(month from age(jobhist.enddate, jobhist.startdate))
        	end
        from emp
        join dept on emp.deptno = dept.deptno
        join jobhist on emp.empno = jobhist.empno;'''

        cur1 = conn.cursor()
        cur1.execute(query1)
        rows1 = cur1.fetchall()  # to execute the query using given cursor

        worksheet2 = workbook.add_worksheet("q2")  # Creating a new Worksheet in given xlsx workbook

        # Creating the Column name
        worksheet2.write('A1', 'Emp Name')
        worksheet2.write('B1', 'Emp No')
        worksheet2.write('C1', 'Dept Name')
        worksheet2.write('D1', 'Total Compensation')
        worksheet2.write('E1', 'Month Spent in Organisation')

        rowIdx = 2

        print()
        print("list_total_compensation")
        print()

        # Writing in xlsx worksheet
        for i in rows1:
            worksheet2.write('A' + str(rowIdx), i[0])
            worksheet2.write('B' + str(rowIdx), i[1])
            worksheet2.write('C' + str(rowIdx), i[2])
            worksheet2.write('D' + str(rowIdx), i[3])
            worksheet2.write('E' + str(rowIdx), i[4])
            rowIdx += 1
            print(i)

        cur1.close()

    # Read and upload the above method xlsx file in postgresDB
    def read_db(self, conn, workbook):
        wb = openpyxl.load_workbook("sqllite1-assignment.xlsx")
        sh1 = wb['q2']  # Reading the q2 worksheet from given Workbook

        row = sh1.max_row
        col = sh1.max_column  # Getting rows and columns from given worksheet

        curr2 = conn.cursor()

        # curr2.execute("insert into postgresDB (ename, empno, dname, sal, month_diff) values (%s, %s, %s, %s, %s)",
        #               ("SMITH", 7369, "RESEARCH", 800, 494))

        # Inserting all the data into postgresDB
        for i in range(2, row + 1):
            li = []
            for j in range(1, col + 1):
                li.append(sh1.cell(i, j).value)
            print(f"#{i}")
            print(li)
            curr2.execute("insert into postgresDB (ename, empno, dname, sal, month_diff) values (%s, %s, %s, %s, %s)",
                          (li[0], li[1], li[2], li[3], li[4]))

        conn.commit()
        curr2.close()
        curx = conn.cursor()
        curx.execute("select * from postgresDB")
        rows = curx.fetchall()
        print(rows)

        curx.close()
    
    # List total compensation given at Department level till date.
    def list_total_compensation_atdept(self, conn, workbook):
        cur3 = conn.cursor()
        
        # Group by dept number to calculate total compensation per dept
        query3 = '''select
        	deptno,
        	dept.dname,
        	sum(sal)
        from postgresDB
        join dept on postgresDB.dname = dept.dname
        group by deptno;'''

        cur3.execute(query3)
        rows2 = cur3.fetchall()  # to execute the query using given cursor

        worksheet3 = workbook.add_worksheet("q4")

        # Writing in xlsx worksheet
        worksheet3.write('A1', 'Dept No')
        worksheet3.write('B1', 'Dept Name')
        worksheet3.write('C1', 'Total Compensation')

        rowIdx = 2

        print()
        print("list_total_compensation_atdept")
        print()

        for r in rows2:
            worksheet3.write('A' + str(rowIdx), r[0])
            worksheet3.write('B' + str(rowIdx), r[1])
            worksheet3.write('C' + str(rowIdx), r[2])
            rowIdx += 1
            print(r)

        cur3.close()


try:
    # Estabilish the DB connection 
    conn = psycopg2.connect(
        user=username,
        host=hostname,
        database=db,
        password=pwd
    )

    workbook = xlsxwriter.Workbook("sqllite1-assignment.xlsx")  # Creating xlsx workbook
    assignment = Assignment()

    assignment.list_emp_details(conn, workbook)  # Calling first method
    assignment.list_total_compensation(conn, workbook)  # Calling Second method
    assignment.read_db(conn, workbook)  # Reading and creating the PostgresDB
    assignment.list_total_compensation_atdept(conn, workbook)
except:
    print("Error in estabilishing the DB connection.")
finally:
    # Closing all the connections
    workbook.close()
    conn.close()
